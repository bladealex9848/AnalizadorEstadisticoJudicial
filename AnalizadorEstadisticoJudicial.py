from rich import print
from rich.table import Table
from rich.console import Console
import os
import glob
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Border, Side
import warnings
import re

console = Console()

def sort_key_func(row):
    """
    Función que devuelve una clave de orden para las filas.
    """
    file_name = row[-1]
    order = ["Primer", "Segundo", "Tercer", "Cuarto"]
    match = re.match(r"(\w+ Trimestre)(_?(\d)?)", file_name)
    if match:
        name, _, number = match.groups()
        return (order.index(name.split(' ')[0]), int(number) if number else 0)
    return (0, 0)

def sorted_files(files):
    """
    Función que ordena los archivos según su orden de trimestre y número.
    """
    return sorted(files, key=sort_key_func)

def log_message(message, add_space=False):
    """
    Función que registra los mensajes tanto en la consola como en un archivo log.
    """
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    log_entry = f"[{timestamp}] {message}"
    with open("log.txt", "a") as log_file:
        if add_space:
            log_file.write("\n")
        log_file.write(log_entry + "\n")

def create_info_table():
    """
    Función que genera y muestra una tabla con información general del programa.
    """
    log_message("Creando tabla de información del programa.")
    table = Table(show_header=True, header_style="bold magenta")
    table.add_column("Versión del script", style="dim", width=20)
    table.add_column("Descripción", style="dim", width=50)
    table.add_column("Nombre del programa", style="dim", width=30)
    table.add_column("Datos propietarios", style="dim", width=30)
    table.add_row("1.1", "Analizador para Organización del Trabajo y Eficiencia en Judicatura", "AnalizadorEstadisticoJudicial", "Consejo Seccional de la Judicatura de Sucre")
    console.print(table)

def create_or_update_readme():
    """
    Función que crea o actualiza el archivo README.md con información sobre el programa y su desarrollador.
    """
    log_message("Creando o actualizando el archivo README.md.")
    contenido_readme = """
# AnalizadorEstadisticoJudicial

## Propósito

Este programa está diseñado para analizar archivos Excel relacionados con la organización del trabajo y la eficiencia en la Judicatura. Automáticamente procesa los datos y genera un consolidado, facilitando la tarea de análisis y presentación de informes.

## ¿Qué hace?

El programa busca archivos Excel con nombres específicos (por ejemplo, 'Primer Trimestre.xls', 'Segundo Trimestre.xls', etc.), incluidos aquellos que tienen un sufijo numérico, como 'Tercer Trimestre_1.xls', en caso de que haya múltiples archivos para un trimestre específico. Procesa cada hoja de estos archivos y finalmente genera un archivo consolidado con los resultados.

## ¿Cómo lo hace?

1. Lee cada archivo Excel.
2. Procesa cada hoja dentro de los archivos.
3. Genera un archivo de resultados por cada archivo Excel.
4. Crea un archivo consolidado con todos los resultados.

## Instrucciones de Uso

1. Asegúrese de que todos los archivos Excel que desea analizar se encuentren en la misma carpeta que este programa.
2. Ejecute el programa.
3. Revise la carpeta 'Consolidado' que se creará automáticamente; allí encontrará los resultados.

Nota: Si ejecuta el programa en la raíz donde están los documentos, se creará automáticamente una carpeta 'Consolidado' con los archivos de resultados.

## Desarrollador

Desarrollado por Alexander Oviedo Fadul, Profesional Universitario Grado 11 del Consejo Seccional de la Judicatura de Sucre.

## Funcionalidades Potenciales

- Integración Directa con SIERJU: Si se obtiene acceso a la API o a alguna interfaz de programación de SIERJU, podríamos adaptar el programa para enviar directamente las estadísticas.
- Análisis Avanzado: Incorporar herramientas de análisis estadístico más avanzadas para obtener insights más detallados de los datos.
- Visualización de Datos: Incorporar gráficos y tableros de mando para una visualización más intuitiva de las estadísticas.

## Adaptabilidad

Es importante que cualquier desarrollo software, especialmente en contextos institucionales, sea fácilmente adaptable a cambios futuros. Esto no solo incluye cambios en la estructura de los datos, sino también en las necesidades del usuario y en el contexto legal y regulatorio.
    """
    with open("README.md", "w") as readme:
        readme.write(contenido_readme)

def create_folder_structure():
    log_message("Creando estructura de carpetas para guardar resultados.")
    if not os.path.exists('Consolidado'):
        os.makedirs('Consolidado')
    subfolder = 'Consolidado/' + datetime.now().strftime('%Y-%m-%d_%H-%M-%S') + '/'
    os.makedirs(subfolder)
    return subfolder

def process_excel_files(excel_files, subfolder):
    log_message("Procesando archivos Excel.")
    all_sheets_data = {}

    # Ordena los archivos por nombre para asegurar que se procesan en el orden correcto
    excel_files.sort()

    for file in excel_files:
        if file not in glob.glob('*Trimestre*.xls*'):
            continue
        log_message(f"Procesando archivo: {file}")
        try:
            xls = pd.ExcelFile(file)
        except Exception as e:
            log_message(f"Error al leer el archivo {file}: {str(e)}")
            console.print(f"[red]Error al leer el archivo {file}: {str(e)}[/red]")
            continue

        result_file = subfolder + file.replace('.xls', '') + '_results.xlsx'
        writer = Workbook()
        writer.remove(writer.active)

        file_table = Table(show_header=True, header_style="bold magenta")
        file_table.add_column("Archivo", style="dim", width=30)
        file_table.add_column("Hoja", style="dim", width=30)
        file_table.add_column("Resultado", style="dim", width=40)

        process_sheets(xls, file, all_sheets_data, writer, file_table)

        try:
            writer.save(result_file)
            log_message(f"Archivo de resultados guardado: {result_file}")
            file_table.add_row(file, "-", f"Archivo de resultados guardado: {result_file}")
        except Exception as e:
            log_message(f"Error al guardar el archivo de resultados: {str(e)}")
            file_table.add_row(file, "-", f"Error al guardar el archivo de resultados: {str(e)}")

        console.print(file_table)

    return all_sheets_data

def process_sheets(xls, file, all_sheets_data, writer, file_table):
    log_message(f"Procesando hojas del archivo: {file}")
    for i, sheet in enumerate(xls.sheet_names):
        log_message(f"Procesando hoja: {sheet}")
        try:
            data = pd.read_excel(file, sheet_name=sheet, header=None)
        except Exception as e:
            log_message(f"Error al leer la hoja {sheet} del archivo {file}: {str(e)}")
            file_table.add_row(file, sheet, f"Error al leer la hoja: {str(e)}")
            continue

        if len(data) >= 20 and 'Total' in data[0].values:
            process_rows(data, file, all_sheets_data, writer, sheet, file_table)
        else:
            log_message(f"La hoja {sheet} del archivo {file} no cumple con las condiciones necesarias")
            file_table.add_row(file, sheet, "La hoja no cumple con las condiciones necesarias")

def process_rows(data, file, all_sheets_data, writer, sheet, file_table):
    log_message(f"Procesando datos de la hoja {sheet} del archivo {file}")
    header_rows = data.iloc[:19].values.tolist()
    row_20_titles = [''] + data.iloc[19, 1:].tolist()
    total_row_index = data[data[0] == 'Total'].index[0]
    total_row_values = ['Total'] + data.iloc[total_row_index, 1:].tolist()

    try:
        # last_column_index = next(i for i, s in enumerate(row_20_titles) if
        #                          "INVENTARIO" in s and ("FINALIZAR" in s or "FINAL" in s)) + 1
        last_column_index = len(row_20_titles)
    except StopIteration:
        last_column_index = len(row_20_titles)

    header_rows = [row[:last_column_index] for row in header_rows]
    row_20_titles = row_20_titles[:last_column_index]
    total_row_values = total_row_values[:last_column_index]

    results_df = pd.DataFrame(header_rows + [row_20_titles, total_row_values])

    ws = writer.create_sheet(title=sheet)
    for row in dataframe_to_rows(results_df, index=False, header=False):
        ws.append(row)

    # Ajustar el texto para los títulos de la fila 19
    for cell in ws[19]:
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

    # Ajustar el texto para las celdas desde la fila 20 hasta el final
    for row in ws.iter_rows(min_row=20, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))

    # Combinar celdas contiguas con el mismo valor en la fila 19
    prev_val = ws.cell(row=19, column=1).value
    merge_start = None
    for col in range(2, ws.max_column + 1):
        current_val = ws.cell(row=19, column=col).value
        if current_val == prev_val:
            if not merge_start:
                merge_start = col - 1
            if col == ws.max_column:
                ws.merge_cells(start_row=19, start_column=merge_start, end_row=19, end_column=col)
        else:
            if merge_start:
                ws.merge_cells(start_row=19, start_column=merge_start, end_row=19, end_column=col - 1)
                merge_start = None
        prev_val = current_val

    if sheet not in all_sheets_data:
        all_sheets_data[sheet] = [row_20_titles]
    all_sheets_data[sheet].append(total_row_values + [file.replace('.xls', '')])

    file_table.add_row(file, sheet, "La hoja ha sido procesada exitosamente")

def consolidate_data(data):
    consolidated_data = [data[0]]  # Títulos
    trimesters = ["Primer Trimestre", "Segundo Trimestre", "Tercer Trimestre", "Cuarto Trimestre"]

    for trimester in trimesters:
        rows = [row for row in data[1:] if trimester in row[-1]]
        if rows:
            consolidated_row = ['Total'] + [sum(row[i] for row in rows) for i in range(1, len(rows[0]) - 1)] + [trimester]
            consolidated_data.append(consolidated_row)

    return consolidated_data

def create_consolidated_file(all_sheets_data, subfolder):
    log_message("Iniciando la creación del archivo consolidado.")
    consolidated_writer = Workbook()
    consolidated_writer.remove(consolidated_writer.active)

    border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))

    for sheet, data in all_sheets_data.items():
        ws = consolidated_writer.create_sheet(title=sheet)

        # Añadir el título (los encabezados de columna)
        column_titles = data[0]  # Esta línea toma la fila de títulos
        ws.append(column_titles)

        # Verificar si necesitamos una segunda tabla
        has_multiple_parts = any("_" in item[-1] for item in data[1:])

        # Primera tabla con datos individuales
        for row in sorted(data[1:], key=sort_key_func):
            ws.append(row)

        # Si hay múltiples partes, agregamos la segunda tabla
        if has_multiple_parts:
            # Agregar un espacio entre las dos tablas
            ws.append([])

            # Segunda tabla
            consolidated_data = consolidate_data(data)
            for row in consolidated_data:
                ws.append(row)

        for row in ws.rows:
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)
                cell.border = border

    consolidated_file = subfolder + 'Consolidado.xlsx'
    try:
        consolidated_writer.save(consolidated_file)
        log_message("Archivo consolidado creado exitosamente en {}.".format(consolidated_file))
        status = "Éxito"
        location = consolidated_file
    except Exception as e:
        log_message("Error al guardar el archivo consolidado: {}.".format(str(e)))
        status = "Error"
        location = "N/A"

    # Mostrar tabla con resultados
    table = Table(show_header=True, header_style="bold magenta")
    table.add_column("Estado", style="dim", width=20)
    table.add_column("Ubicación", style="dim", width=60)
    table.add_row(status, location)
    console.print(table)

def main():
    try:
        warnings.filterwarnings('ignore', category=UserWarning, module='xlrd')

        # Verifica si existe log.txt, si no, lo crea
        if not os.path.exists("log.txt"):
            with open("log.txt", "w") as log_file:
                log_file.write("Registro para AnalizadorEstadisticoJudicial\n\n")

        # Verifica si existe README.md, si no, lo crea
        if not os.path.exists("README.md"):
            create_or_update_readme()

        create_info_table()

        log_message("Iniciando el procesamiento de los archivos Excel.", add_space=True)

        subfolder = create_folder_structure()

        # Obtenemos una lista de todos los archivos .xls en el directorio actual
        all_files = glob.glob('* Trimestre*.xls')

        # Ordenamos los archivos antes de procesarlos
        excel_files = sorted_files(all_files)

        all_sheets_data = process_excel_files(excel_files, subfolder)
        create_consolidated_file(all_sheets_data, subfolder)

        log_message("Procesamiento finalizado.", add_space=True)

        input("\nPresiona cualquier tecla para salir...")
    except Exception as e:
        log_message(f"Error inesperado: {str(e)}")
        console.print(
            f"[red]Error inesperado: {str(e)}. Por favor, consulte el archivo log.txt para obtener más detalles.[/red]")

if __name__ == "__main__":
    main()